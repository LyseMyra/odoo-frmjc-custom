import base64
import io
import logging

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


class TrainingRubanImportWizard(models.TransientModel):
    _name = 'training.ruban.import.wizard'
    _description = "Assistant d'import du ruban pédagogique depuis Excel"

    state = fields.Selection([
        ('config', 'Configuration'),
        ('done', 'Résultat'),
    ], default='config')

    # ── Paramètres principaux ──────────────────────────────────────
    formation_id = fields.Many2one(
        'training.formation', string='Formation', required=True,
    )
    fichier = fields.Binary(string='Fichier Excel (.xlsx)', required=True)
    fichier_nom = fields.Char()
    ligne_debut = fields.Integer(
        string='1re ligne de données',
        default=3,
        help='Numéro de la 1re ligne contenant des données (par ex. 3 si les 2 premières lignes sont des en-têtes).',
    )
    ecraser_modules = fields.Boolean(
        string='Mettre à jour les modules existants',
        default=False,
        help='Si coché, les champs texte des modules déjà présents seront mis à jour. '
             'Les entrées de répartition existantes ne sont jamais supprimées.',
    )

    # ── Configuration avancée des colonnes ────────────────────────
    col_bloc     = fields.Integer('BLOC',             default=1)
    col_module   = fields.Integer('Module (intitulé)', default=2)
    col_objectifs= fields.Integer('Objectifs',         default=3)
    col_contenus = fields.Integer('Contenus',         default=4)
    col_interv   = fields.Integer('Intervenant',      default=5)
    col_peda     = fields.Integer('Modalités péda',   default=6)
    col_sem      = fields.Integer('N° semaine',       default=7)
    col_vol_of   = fields.Integer('Volume OF (h)',    default=9)
    col_act      = fields.Integer('Activités ENT',    default=10)
    col_cmd      = fields.Integer('Commandes MA',     default=11)
    col_vol_ent  = fields.Integer('Volume ENT (h)',   default=12)
    col_lundi    = fields.Integer('Lundi (h)',        default=14)
    col_mardi    = fields.Integer('Mardi (h)',        default=15)
    col_mercredi = fields.Integer('Mercredi (h)',     default=16)
    col_jeudi    = fields.Integer('Jeudi (h)',        default=17)
    col_vendredi = fields.Integer('Vendredi (h)',     default=18)

    # ── Résultat ──────────────────────────────────────────────────
    rapport = fields.Text(readonly=True)
    nb_blocs   = fields.Integer(readonly=True)
    nb_modules = fields.Integer(readonly=True)
    nb_reps    = fields.Integer(readonly=True)
    nb_alertes = fields.Integer(readonly=True)

    # ── Action principale ──────────────────────────────────────────
    def action_importer(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise UserError(
                "La bibliothèque openpyxl est requise.\n"
                "Installez-la avec :  pip install openpyxl"
            )

        if not self.fichier:
            raise UserError("Veuillez sélectionner un fichier Excel (.xlsx).")

        data = base64.b64decode(self.fichier)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise UserError(f"Impossible d'ouvrir le fichier Excel : {e}")

        ws = wb.active

        # ── Résoudre les cellules fusionnées ──────────────────────
        # openpyxl met None dans les cellules non-maîtresses d'un merge.
        # On recrée un dict {(row, col): valeur_de_la_cellule_maîtresse}.
        merged_values = {}
        for merge_range in ws.merged_cells.ranges:
            master_val = ws.cell(merge_range.min_row, merge_range.min_col).value
            for r in range(merge_range.min_row, merge_range.max_row + 1):
                for c in range(merge_range.min_col, merge_range.max_col + 1):
                    merged_values[(r, c)] = master_val

        def cv(row, col):
            """Valeur de la cellule (gère les fusions), toujours str stripped."""
            key = (row, col)
            raw = merged_values.get(key, ws.cell(row, col).value)
            if raw is None:
                return ''
            return str(raw).strip()

        def fv(row, col):
            """Valeur float d'une cellule numérique (0.0 si vide/invalide)."""
            raw = merged_values.get((row, col), ws.cell(row, col).value)
            if raw is None or raw == '':
                return 0.0
            try:
                return float(raw)
            except (ValueError, TypeError):
                return 0.0

        # ── Correspondance jour → colonne ──────────────────────────
        jours = {
            'lundi':    self.col_lundi,
            'mardi':    self.col_mardi,
            'mercredi': self.col_mercredi,
            'jeudi':    self.col_jeudi,
            'vendredi': self.col_vendredi,
        }

        formation = self.formation_id
        blocs_cache   = {}   # (formation_id, bloc_name)  → training.bloc
        modules_cache = {}   # (bloc_id, module_code)     → training.module

        log = []
        nb_blocs = nb_modules = nb_reps = nb_alertes = 0
        current_bloc_raw = ''

        max_row = ws.max_row or 0

        for row in range(self.ligne_debut, max_row + 1):

            # ── Ignorer les lignes entièrement vides ──────────────
            if all(ws.cell(row, c).value is None for c in range(1, 20)):
                continue

            # ── BLOC : valeur persistée si cellule fusionnée ──────
            bloc_val = cv(row, self.col_bloc)
            if bloc_val:
                current_bloc_raw = bloc_val

            # ── Module intitulé ───────────────────────────────────
            module_raw = cv(row, self.col_module)
            if not module_raw:
                continue

            intitule = module_raw.strip()

            # ── Bloc par défaut (BAT) si aucun bloc renseigné ────
            if not current_bloc_raw:
                current_bloc_raw = 'BAT'

            # ── Colonnes texte ────────────────────────────────────
            objectifs     = cv(row, self.col_objectifs)
            contenus_raw  = cv(row, self.col_contenus)
            interv_raw    = cv(row, self.col_interv)
            peda_raw      = cv(row, self.col_peda)
            act_raw       = cv(row, self.col_act)
            cmd_raw       = cv(row, self.col_cmd)
            vol_of        = fv(row, self.col_vol_of)
            vol_ent       = fv(row, self.col_vol_ent)
            semaine_raw   = cv(row, self.col_sem)

            # ── Trouver / créer le BLOC ───────────────────────────
            bloc_name = current_bloc_raw.replace(' ', '')  # "BC 2" → "BC2"
            bloc_key = (formation.id, bloc_name)

            if bloc_key not in blocs_cache:
                type_bloc = 'bat' if bloc_name.upper().startswith('BAT') else 'bc'
                if bloc_name.upper().startswith('BC') and '-' in bloc_name:
                    type_bloc = 'transversal'

                bloc = self.env['training.bloc'].search([
                    ('formation_id', '=', formation.id),
                    ('name', '=', bloc_name),
                ], limit=1)
                if not bloc:
                    bloc = self.env['training.bloc'].create({
                        'formation_id': formation.id,
                        'name': bloc_name,
                        'type_bloc': type_bloc,
                    })
                    nb_blocs += 1
                    log.append(f'[BLOC créé] {bloc_name}')
                blocs_cache[bloc_key] = bloc

            bloc = blocs_cache[bloc_key]

            # ── Trouver / créer le MODULE ─────────────────────────
            module_key = (bloc.id, intitule)

            if module_key not in modules_cache:
                module_r = self.env['training.module'].search([
                    ('bloc_id', '=', bloc.id),
                    ('intitule', '=', intitule),
                ], limit=1)

                if not module_r:
                    module_r = self.env['training.module'].create({
                        'bloc_id':              bloc.id,
                        'intitule':             intitule,
                        'objectifs':            objectifs,
                        'contenus':             contenus_raw,
                        'modalites_peda':       peda_raw,
                        'activites_entreprise': act_raw,
                        'commandes_tuteur':     cmd_raw,
                        'volume_of':            vol_of,
                        'volume_entreprise':    vol_ent,
                    })
                    nb_modules += 1
                    log.append(f'  [MODULE créé] {bloc_name} / {intitule}')

                    # Intervenant par défaut
                    if interv_raw:
                        interv = self.env['res.partner'].search([
                            ('name', 'ilike', interv_raw),
                            ('is_company', '=', False),
                        ], limit=1)
                        if interv:
                            module_r.intervenant_default_id = interv.id
                        else:
                            nb_alertes += 1
                            log.append(f'    ⚠ Intervenant introuvable : {interv_raw!r}')

                elif self.ecraser_modules:
                    module_r.write({
                        'objectifs':            objectifs or module_r.objectifs,
                        'contenus':             contenus_raw or module_r.contenus,
                        'modalites_peda':       peda_raw or module_r.modalites_peda,
                        'activites_entreprise': act_raw or module_r.activites_entreprise,
                        'commandes_tuteur':     cmd_raw or module_r.commandes_tuteur,
                        'volume_of':            vol_of or module_r.volume_of,
                        'volume_entreprise':    vol_ent or module_r.volume_entreprise,
                    })
                    log.append(f'  [MODULE màj] {bloc_name} / {intitule}')

                modules_cache[module_key] = module_r

            module_r = modules_cache[module_key]

            # ── Répartition : une entrée par jour avec heures > 0 ─
            try:
                semaine = int(float(semaine_raw)) if semaine_raw else 0
            except (ValueError, TypeError):
                semaine = 0

            if semaine <= 0:
                continue

            for jour, col_idx in jours.items():
                heures = fv(row, col_idx)
                if heures <= 0:
                    continue

                already = self.env['training.module.repartition'].search_count([
                    ('module_id', '=', module_r.id),
                    ('semaine', '=', semaine),
                    ('jour', '=', jour),
                ])
                if already:
                    continue

                self.env['training.module.repartition'].create({
                    'module_id': module_r.id,
                    'semaine':   semaine,
                    'jour':      jour,
                    'heures':    heures,
                })
                nb_reps += 1

        # ── Résumé ────────────────────────────────────────────────
        summary = [
            f'Formation : {formation.display_name}',
            f'Blocs créés      : {nb_blocs}',
            f'Modules créés    : {nb_modules}',
            f'Répartitions ajoutées : {nb_reps}',
        ]
        if nb_alertes:
            summary.append(f'Alertes          : {nb_alertes} (voir détail ci-dessous)')
        if log:
            summary += ['', '── Détail ──'] + log

        self.write({
            'state':      'done',
            'rapport':    '\n'.join(summary),
            'nb_blocs':   nb_blocs,
            'nb_modules': nb_modules,
            'nb_reps':    nb_reps,
            'nb_alertes': nb_alertes,
        })

        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }

    def action_retour(self):
        self.state = 'config'
        return {
            'type':      'ir.actions.act_window',
            'res_model': self._name,
            'res_id':    self.id,
            'view_mode': 'form',
            'target':    'new',
        }
