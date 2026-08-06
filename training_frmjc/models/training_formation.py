import base64
import io

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError


class TrainingFormation(models.Model):
    _name = 'training.formation'
    _description = 'Contenu de la formation (référentiel stable)'
    _order = 'intitule'
    _rec_name = 'intitule'

    # ── Identification ─────────────────────────────────────────────
    intitule = fields.Char(
        string='Intitulé',
        required=True,
    )
    type_formation = fields.Selection(
        selection=[
            ('dejeps', 'DEJEPS'),
            ('desjeps', 'DESJEPS'),
            ('bpjeps', 'BPJEPS'),
            ('uniformation', 'Uniformation'),
            ('interne', 'Formation interne'),
        ],
        string='Type',
        required=True,
    )
    specialite = fields.Char(string='Spécialité')
    mention = fields.Char(string='Mention')

    # ── Volumes horaires ───────────────────────────────────────────
    volume_centre = fields.Float(
        string='Volume horaire en centre (h)',
        required=True,
        help='Total des heures en organisme de formation'
    )
    volume_entreprise = fields.Float(
        string='Volume horaire en entreprise (h)',
        required=True,
    )
    nb_stagiaires_max = fields.Integer(
        string='Nombre maximum de stagiaires',
        required=True,
        default=20,
    )
    tarif_horaire = fields.Float(
        string='Tarif horaire (€/h)',
        digits=(10, 2),
        help='Montant facturé par heure de formation (ex : 15,00 € pour DEJEPS). '
             'Sert au calcul des recettes théoriques et réelles par stagiaire.',
    )

    # ── Informations réglementaires ────────────────────────────────
    reference_arrete = fields.Char(
        string='Référence de l\'arrêté de création',
    )

    # ── Relations ──────────────────────────────────────────────────
    bloc_ids = fields.One2many(
        'training.bloc',
        'formation_id',
        string='Blocs pédagogiques (BC / BAT)',
    )
    nb_modules = fields.Integer(
        string='Nombre de modules',
        compute='_compute_nb_modules',
    )
    session_ids = fields.One2many(
        'training.session',
        'formation_id',
        string='Sessions',
    )

    # ── Contraintes SQL ────────────────────────────────────────────
    _unique_formation = models.Constraint(
        'UNIQUE(intitule, type_formation, specialite, mention)',
        'Une formation avec ces caractéristiques existe déjà.'
    )
    _check_nb_stagiaires = models.Constraint(
        'CHECK(nb_stagiaires_max > 0)',
        'Le nombre maximum de stagiaires doit être positif.'
    )
    
    # ── Méthodes compute ───────────────────────────────────────────
    @api.depends('bloc_ids.module_ids')
    def _compute_nb_modules(self):
        for rec in self:
            rec.nb_modules = sum(len(b.module_ids) for b in rec.bloc_ids)

    # ── Contraintes Python ─────────────────────────────────────────
    @api.constrains('volume_centre', 'volume_entreprise')
    def _check_volumes(self):
        for rec in self:
            if rec.volume_centre <= 0:
                raise ValidationError(
                    'Le volume horaire en centre doit être positif.'
                )
            if rec.volume_entreprise < 0:
                raise ValidationError(
                    'Le volume horaire en entreprise ne peut pas être négatif.'
                )

    # ── Actions ────────────────────────────────────────────────────
    def action_ouvrir_import_ruban(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'training.ruban.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_formation_id': self.id},
        }

    def action_exporter_ruban_excel(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ruban pédagogique"

        # ── Styles ──────────────────────────────────────────────────
        thin = Side(style='thin', color='AAAAAA')
        thick = Side(style='medium', color='444444')
        border_data = Border(left=thin, right=thin, top=thin, bottom=thin)
        border_bloc = Border(left=thick, right=thick, top=thick, bottom=thick)
        c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        l_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

        fill_title  = PatternFill(fill_type='solid', fgColor='1F4E79')
        fill_header = PatternFill(fill_type='solid', fgColor='2E75B6')
        fill_bloc   = PatternFill(fill_type='solid', fgColor='D6E4F7')
        fill_module = PatternFill(fill_type='solid', fgColor='EBF3FB')

        font_title  = Font(bold=True, size=12, color='FFFFFF')
        font_header = Font(bold=True, size=9,  color='FFFFFF')
        font_bloc   = Font(bold=True, size=9,  color='1F4E79')
        font_data   = Font(size=9)

        # ── Largeurs de colonnes ─────────────────────────────────────
        col_widths = {
            1: 8, 2: 10, 3: 32, 4: 32, 5: 20, 6: 25,
            7: 7, 8: 1, 9: 7, 10: 22, 11: 22, 12: 7, 13: 7,
            14: 4, 15: 4, 16: 4, 17: 4, 18: 4,
        }
        for col, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # ── Ligne 1 : titre ──────────────────────────────────────────
        ws.merge_cells('A1:R1')
        c = ws['A1']
        c.value     = f"RUBAN DE CONTRÔLES — {self.display_name.upper()}"
        c.font      = font_title
        c.fill      = fill_title
        c.alignment = c_align
        ws.row_dimensions[1].height = 22

        # ── Ligne 2 : en-têtes de colonnes ──────────────────────────
        HEADERS = [
            (1, 'BLOC'),          (2, 'Module\n(intitulé)'),
            (3, 'Objectifs'),     (4, 'Contenus'),
            (5, 'Intervenant'),   (6, 'Modalités\npédagogiques'),
            (7, 'N°\nsem.'),      (8, ''),
            (9, 'Vol.\nOF (h)'),  (10, 'Activités\nENT'),
            (11, 'Cmdes\nMA'),    (12, 'Vol.\nENT (h)'),
            (13, 'Total\n(h)'),
            (14, 'L'), (15, 'M'), (16, 'Me'), (17, 'J'), (18, 'V'),
        ]
        for col, label in HEADERS:
            c = ws.cell(2, col, label)
            c.font      = font_header
            c.fill      = fill_header
            c.alignment = c_align
            c.border    = border_data
        ws.row_dimensions[2].height = 30

        # ── Gel des volets ──────────────────────────────────────────
        ws.freeze_panes = 'A3'

        # ── Données ─────────────────────────────────────────────────
        JOURS = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi']
        DAY_COL = {j: 14 + i for i, j in enumerate(JOURS)}

        def _set(row, col, value, align=None):
            c = ws.cell(row, col, value)
            c.font   = font_data
            c.border = border_data
            c.alignment = align or l_align
            return c

        def _merge_range(r1, c1, r2, c2, value, fill=None, font=None, align=None):
            ws.merge_cells(start_row=r1, start_column=c1,
                           end_row=r2,   end_column=c2)
            c = ws.cell(r1, c1, value)
            c.font      = font  or font_data
            c.border    = border_data
            c.alignment = align or l_align
            if fill:
                c.fill = fill
            # border sur toutes les cellules de la fusion
            for rr in range(r1, r2 + 1):
                for cc in range(c1, c2 + 1):
                    ws.cell(rr, cc).border = border_data

        current_row = 3

        for bloc in self.bloc_ids.sorted('sequence'):
            modules = bloc.module_ids.sorted('sequence')
            if not modules:
                continue

            bloc_row_start = current_row

            for module in modules:
                mod_row_start = current_row

                # Semaines présentes pour ce module (triées)
                reps = module.repartition_ids.sorted('semaine')
                semaines = sorted(set(r.semaine for r in reps)) if reps else [0]

                for semaine in semaines:
                    # Colonnes semaine + jours (une ligne par semaine)
                    if semaine > 0:
                        _set(current_row, 7, semaine, align=c_align)
                    for jour in JOURS:
                        week_reps = reps.filtered(
                            lambda r, s=semaine, j=jour: r.semaine == s and r.jour == j
                        )
                        if week_reps:
                            _set(current_row, DAY_COL[jour],
                                 week_reps[0].heures, align=c_align)
                        else:
                            _set(current_row, DAY_COL[jour], None, align=c_align)

                    # Col H toujours vide
                    _set(current_row, 8, None, align=c_align)
                    current_row += 1

                mod_row_end = current_row - 1

                interv_name = (module.intervenant_default_id.name
                               if module.intervenant_default_id else '')

                # Colonnes module fusionnées sur toutes ses semaines
                module_cols = [
                    (2,  module.intitule or '',           l_align, fill_module, font_bloc),
                    (3,  module.objectifs or '',          l_align, fill_module, None),
                    (4,  module.contenus or '',            l_align, fill_module, None),
                    (5,  interv_name,                     c_align, fill_module, None),
                    (6,  module.modalites_peda or '',     l_align, fill_module, None),
                    (9,  module.volume_of or 0,           c_align, fill_module, None),
                    (10, module.activites_entreprise or '',l_align, fill_module, None),
                    (11, module.commandes_tuteur or '',   l_align, fill_module, None),
                    (12, module.volume_entreprise or 0,  c_align, fill_module, None),
                    (13, module.volume_total or 0,        c_align, fill_module, None),
                ]
                for col, val, al, fl, fo in module_cols:
                    if mod_row_end > mod_row_start:
                        _merge_range(mod_row_start, col, mod_row_end, col,
                                     val, fill=fl,
                                     font=fo or font_data, align=al)
                    else:
                        c = ws.cell(mod_row_start, col, val)
                        c.font      = fo or font_data
                        c.fill      = fl
                        c.border    = border_data
                        c.alignment = al

            bloc_row_end = current_row - 1

            # Colonne BLOC fusionnée
            if bloc_row_end > bloc_row_start:
                _merge_range(bloc_row_start, 1, bloc_row_end, 1,
                             bloc.name, fill=fill_bloc,
                             font=font_bloc, align=c_align)
            else:
                c = ws.cell(bloc_row_start, 1, bloc.name)
                c.font      = font_bloc
                c.fill      = fill_bloc
                c.border    = border_data
                c.alignment = c_align

        # ── Sauvegarde et téléchargement ─────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = (self.intitule or 'formation').replace(' ', '_')[:40]
        filename = f"ruban_{safe_name}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name':      filename,
            'datas':     base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id':    self.id,
            'mimetype':  ('application/vnd.openxmlformats-'
                          'officedocument.spreadsheetml.sheet'),
        })

        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{attachment.id}/{filename}?download=true',
            'target': 'new',
        }

    # ── Représentation ─────────────────────────────────────────────
    def _compute_display_name(self):
        for rec in self:
            name = rec.intitule or '?'
            if rec.mention:
                name += f' - {rec.mention}'
            if rec.specialite:
                name += f' ({rec.specialite})'
            rec.display_name = name