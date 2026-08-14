import base64
import datetime
import io
import logging

from odoo import models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# (libellé, clé de champ sur mobility.mobility, type) — un champ par ligne
# du fichier (fichier vertical Champ/Valeur, individuel à une mobilité,
# pas un tableau multi-lignes). Chaque field_key est un champ réel déjà
# présent sur mobility.mobility (mêmes noms que le formulaire public) :
# la pré-remplissage à l'export se fait par simple getattr.
FICHE_FIELDS = [
    ('Prénom du volontaire', 'participant_prenom', 'char'),
    ('Nom du volontaire', 'participant_nom', 'char'),
    ('Email du volontaire', 'email', 'char'),
    ('Téléphone', 'telephone', 'char'),
    ('Date de naissance', 'date_naissance', 'date'),
    ('Nationalité (pays)', 'nationalite_id', 'country'),
    ('Pays de résidence', 'pays_residence_id', 'country'),
    ('Ville de résidence', 'ville_residence', 'char'),
    ("N° pièce d'identité", 'id_document_numero', 'char'),
    ("Validité pièce d'identité", 'id_document_validite', 'date'),
    ('Besoins particuliers', 'besoins_particuliers', 'char'),
    ('Contact urgence - Nom', 'contact_urgence_nom', 'char'),
    ('Contact urgence - Lien', 'contact_urgence_lien', 'char'),
    ('Contact urgence - Adresse', 'contact_urgence_adresse', 'char'),
    ('Contact urgence - Téléphone', 'contact_urgence_telephone', 'char'),
    ('Contact urgence - Email', 'contact_urgence_email', 'char'),
    ('Pays de mission', 'country_id', 'country'),
    ('Ville de mission', 'city', 'char'),
    ("ID de l'offre", 'offer_code', 'char'),
    ("Titre de l'offre", 'titre_offre', 'char'),
    ('Date de publication', 'date_publication_offre', 'date'),
    ('Date de sélection', 'date_selection_offre', 'date'),
    ('Date de début proposée', 'date_debut_proposee', 'date'),
    ('Date de fin proposée', 'date_fin_proposee', 'date'),
    ("Source de l'offre", 'source_offre', 'char'),
    ("Structure d'envoi - Nom", 'sending_org_nom', 'char'),
    ("Structure d'envoi - OID", 'sending_org_oid_saisi', 'char'),
    ("Structure d'envoi - Nom légal", 'sending_org_nom_legal', 'char'),
    ("Structure d'envoi - Adresse", 'sending_org_adresse', 'char'),
    ("Structure d'envoi - Type d'organisme", 'sending_org_type_organisme', 'char'),
    ("Structure d'envoi - Email", 'sending_org_email', 'char'),
    ("Structure d'envoi - Téléphone", 'sending_org_telephone', 'char'),
    ("Structure d'envoi - Coordinateur", 'sending_contact_nom', 'char'),
    ("Structure d'envoi - Email coordinateur", 'sending_contact_email', 'char'),
    ("Structure d'envoi - Téléphone coordinateur", 'sending_contact_telephone', 'char'),
    ("Structure d'accueil - Nom", 'hosting_org_nom', 'char'),
    ("Structure d'accueil - OID", 'hosting_org_oid_saisi', 'char'),
    ("Structure d'accueil - Nom légal", 'hosting_org_nom_legal', 'char'),
    ("Structure d'accueil - Adresse", 'hosting_org_adresse', 'char'),
    ("Structure d'accueil - Type d'organisme", 'hosting_org_type_organisme', 'char'),
    ("Structure d'accueil - Email", 'hosting_org_email', 'char'),
    ("Structure d'accueil - Téléphone", 'hosting_org_telephone', 'char'),
    ("Structure d'accueil - Référent", 'hosting_contact_nom', 'char'),
    ("Structure d'accueil - Email référent", 'hosting_contact_email', 'char'),
    ("Structure d'accueil - Téléphone référent", 'hosting_contact_telephone', 'char'),
    ('Déclaration acceptée (Oui/Non)', 'declaration_acceptee', 'bool'),
    ('Date de déclaration', 'date_declaration', 'date'),
    ('Lieu de déclaration', 'lieu_declaration', 'char'),
    ('Signature (nom saisi)', 'signature_nom', 'char'),
]

FORMAT_HINTS = {
    'date': 'JJ/MM/AAAA',
    'country': "Nom du pays tel qu'enregistré dans Odoo (ex. France, Allemagne...).",
    'bool': 'Oui / Non',
    'char': 'Texte libre',
}

# Précisions par champ, prioritaires sur le hint générique par type — utile
# quand un champ a un effet moins évident (ex. l'âge du participant a
# besoin d'une date de référence en plus de la date de naissance).
FIELD_HINT_OVERRIDES = {
    'date_naissance': (
        "JJ/MM/AAAA — combinée à la date de début (réelle si connue, sinon "
        "proposée ci-dessous), sert à calculer l'âge du participant."
    ),
    'date_debut_proposee': (
        "JJ/MM/AAAA — sert aussi de date de référence pour calculer l'âge "
        "du participant tant que la date de début réelle n'est pas connue."
    ),
}

REF_LABEL = 'Référence mobilité (ne pas modifier)'

# Mêmes champs obligatoires que le formulaire public (cf. champs_requis
# dans portal_fiche_renseignement.py) — un fichier incomplet est rejeté
# plutôt qu'écrit partiellement en silence (ex. sans date de référence,
# l'âge du participant ne peut pas être calculé et reste à 0).
CHAMPS_REQUIS = [
    ('participant_prenom', 'Prénom du volontaire'),
    ('participant_nom', 'Nom du volontaire'),
    ('email', 'Email du volontaire'),
    ('date_naissance', 'Date de naissance'),
    ('hosting_org_nom', "Nom de la structure d'accueil"),
    ('declaration_acceptee', "Déclaration acceptée (Oui/Non)"),
]


def _parse_date_cell(value):
    if not value:
        return False
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    return False


def _parse_bool_cell(value):
    return str(value).strip().lower() in ('1', 'true', 'vrai', 'oui', 'yes', 'x')


class MobilityImportFicheWizard(models.TransientModel):
    _name = 'mobility.import.fiche.wizard'
    _description = "Assistant export/import Excel — fiche de renseignement d'une mobilité"

    mobility_id = fields.Many2one(
        'mobility.mobility', string='Mobilité', required=True,
        default=lambda self: self.env.context.get('active_id'),
    )
    file = fields.Binary(string='Fichier Excel (.xlsx)')
    filename = fields.Char()

    def action_telecharger_modele(self):
        """Génère un fichier Excel individuel à cette mobilité : une ligne
        par champ (mise en page verticale Champ/Valeur/Aide), pré-rempli
        avec les valeurs déjà connues — utilisable aussi bien comme
        modèle vierge que comme relevé à compléter/corriger."""
        self.ensure_one()
        if not self.mobility_id:
            raise UserError('Aucune mobilité sélectionnée.')
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. pip install openpyxl")

        mobility = self.mobility_id
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fiche renseignement'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        for col_idx, label in enumerate(['Champ', 'Valeur', 'Format / aide'], start=1):
            cell = ws.cell(1, col_idx, label)
            cell.font = header_font
            cell.fill = header_fill
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 55
        ws.freeze_panes = 'A2'

        ws.append([
            REF_LABEL, mobility.name,
            "Sert à vérifier la correspondance à l'import — ne pas modifier.",
        ])
        ref_fill = PatternFill(fill_type='solid', fgColor='D9D9D9')
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = ref_fill

        for label, field_key, kind in FICHE_FIELDS:
            value = mobility[field_key]
            if kind == 'country':
                display = value.name if value else ''
            elif kind == 'bool':
                display = 'Oui' if value else 'Non'
            else:
                display = value or ''
            hint = FIELD_HINT_OVERRIDES.get(field_key, FORMAT_HINTS[kind])
            ws.append([label, display, hint])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"fiche_{mobility.name.replace('/', '-')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(output.read()),
            'res_model': 'mobility.import.fiche.wizard',
            'res_id': self.id,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}/{filename}?download=true',
            'target': 'new',
        }

    def action_importer(self):
        """Relit le fichier individuel rempli et écrit les informations de
        la fiche de renseignement (§4) sur CETTE mobilité (self.mobility_id).
        La ligne « Référence mobilité » du fichier est vérifiée contre la
        mobilité courante — importer le fichier d'une autre mobilité lève
        une erreur plutôt que d'écraser silencieusement la mauvaise fiche."""
        self.ensure_one()
        if not self.mobility_id:
            raise UserError('Aucune mobilité sélectionnée.')
        if not self.file:
            raise UserError('Sélectionnez un fichier à importer.')
        if self.mobility_id.fiche_renseignement_validee:
            raise UserError(
                'La fiche de renseignement de cette mobilité est déjà '
                'validée — repassez-la en non-validée avant de réimporter.'
            )
        try:
            import openpyxl
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. pip install openpyxl")

        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(self.file)), data_only=True,
        )
        ws = wb['Fiche renseignement'] if 'Fiche renseignement' in wb.sheetnames else wb.active

        label_map = {label: (field_key, kind) for label, field_key, kind in FICHE_FIELDS}
        values_by_label = {}
        ref_value = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            value = row[1] if len(row) > 1 else None
            if label == REF_LABEL:
                ref_value = value
            elif label in label_map:
                values_by_label[label] = value

        if ref_value and str(ref_value).strip() != self.mobility_id.name:
            raise UserError(
                f'Ce fichier correspond à la mobilité « {ref_value} », pas à '
                f'« {self.mobility_id.name} » — vérifiez le fichier importé.'
            )

        Country = self.env['res.country']
        country_cache = {}

        def _resolve_country(name):
            if not name:
                return False
            key = str(name).strip().lower()
            if key not in country_cache:
                country_cache[key] = Country.search([('name', '=ilike', key)], limit=1)
            return country_cache[key]

        post = {}
        for label, (field_key, kind) in label_map.items():
            value = values_by_label.get(label)
            if kind == 'date':
                post[field_key] = _parse_date_cell(value)
            elif kind == 'country':
                country = _resolve_country(value)
                post[field_key] = country.id if country else False
            elif kind == 'bool':
                post[field_key] = _parse_bool_cell(value)
            else:
                post[field_key] = value

        manquants = [label for key, label in CHAMPS_REQUIS if not post.get(key)]
        if manquants:
            raise UserError(
                "Fichier incomplet — champs obligatoires manquants :\n"
                + '\n'.join(f'- {label}' for label in manquants)
                + "\n\n(mêmes champs requis que sur le formulaire public)"
            )

        vals = self.env['mobility.mobility']._build_fiche_vals(post)
        self.mobility_id.write(vals)
        self.env['mobility.document'].create({
            'mobility_id': self.mobility_id.id,
            'participant_id': self.mobility_id.participant_id.id,
            'document_type': 'fiche_renseignement',
            'statut': 'valide',
            'upload_date': fields.Date.today(),
            'emis_par': 'Import Excel (secrétariat)',
            'notes': (
                "Fiche de renseignement importée depuis un fichier Excel "
                "par le secrétariat, sans passage par le formulaire public."
            ),
        })
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import terminé',
                'message': f'Fiche de renseignement mise à jour pour {self.mobility_id.name}.',
                'type': 'success',
                'sticky': False,
            },
        }
