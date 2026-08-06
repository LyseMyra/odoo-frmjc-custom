import base64
import io

from odoo import models, fields, api
from odoo.exceptions import UserError


class MobilityExportBmWizard(models.TransientModel):
    _name = 'mobility.export.bm.wizard'
    _description = 'Assistant export BM (Beneficiary Module)'

    grant_id = fields.Many2one('mobility.grant', string='Convention de subvention')
    mobility_ids = fields.Many2many('mobility.mobility', string='Mobilités à exporter')

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        active_model = self.env.context.get('active_model')
        active_ids = self.env.context.get('active_ids') or []
        if active_model == 'mobility.grant' and active_ids:
            grant = self.env['mobility.grant'].browse(active_ids[0])
            res['grant_id'] = grant.id
            res['mobility_ids'] = [(6, 0, grant.mobility_ids.ids)]
        elif active_model == 'mobility.mobility' and active_ids:
            res['mobility_ids'] = [(6, 0, active_ids)]
        return res

    def action_exporter(self):
        self.ensure_one()
        if not self.mobility_ids:
            raise UserError('Sélectionnez au moins une mobilité à exporter.')

        # ── Contrôle préalable obligatoire (§12 du cahier) ──────────────
        erreurs = []
        for mobility in self.mobility_ids:
            erreurs += mobility._check_bm_export_ready()
        if erreurs:
            raise UserError(
                "Export bloqué — champs obligatoires manquants :\n\n"
                + '\n'.join(erreurs)
            )

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. pip install openpyxl")

        columns = self._get_bm_columns()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Export BM'

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        for col_idx, (label, _getter) in enumerate(columns, start=1):
            cell = ws.cell(1, col_idx, label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = 'A2'

        for row_idx, mobility in enumerate(self.mobility_ids, start=2):
            for col_idx, (_label, getter) in enumerate(columns, start=1):
                ws.cell(row_idx, col_idx, getter(mobility))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"export_bm_{fields.Date.today().isoformat()}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': base64.b64encode(output.read()),
            'res_model': 'mobility.export.bm.wizard',
            'res_id': self.id,
            'mimetype': (
                'application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sheet'
            ),
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}/{filename}?download=true',
            'target': 'new',
        }

    def _get_bm_columns(self):
        """Liste (libellé, fonction(mobility) -> valeur) des colonnes
        exportées, dans l'ordre reconstruit à partir du cahier des
        charges (§9-§13) et de l'échantillon BM réel fourni.

        ⚠️ Reconstruction BEST-EFFORT, à vérifier contre un modèle BM
        officiel propre avant tout envoi réel : l'échantillon disponible
        est un export de données (PDF/OCR), pas le gabarit de colonnes
        lui-même — son en-tête était disloqué à l'extraction. Couverture
        partielle : les ~20 champs d'adresse/légal détaillés par
        structure (§8) ne sont pas encore mappés colonne par colonne.
        """

        def _label(field_name, value):
            selection = dict(self.env['mobility.mobility']._fields[field_name].selection)
            return selection.get(value, '')

        def _oui_non(value):
            return 'YES' if value else 'NO'

        def _date(value):
            return value.strftime('%d/%m/%Y') if value else ''

        def _soutien_partner(m):
            # L'échantillon BM réel (mobilités Accueil) montre que
            # « Organisme de soutien » désigne la structure d'envoi ;
            # le cahier (§3) indique qu'en Envoi c'est l'organisme FR
            # (modélisé ici par support_partner_id). D'où la bascule.
            return (
                m.sending_partner_id if m.mobility_direction == 'accueil'
                else m.support_partner_id
            )

        return [
            ('Numéro de convention de subvention',
             lambda m: m.grant_id.numero_convention or ''),
            ('Habilitation label LEAD',
             lambda m: m.habilitation_id.numero_habilitation or ''),
            ('Programme', lambda m: _label('programme', m.programme)),
            ('Type mobilité', lambda m: _label('mobility_direction', m.mobility_direction)),
            ('Durée mobilité', lambda m: _label('mobility_duration', m.mobility_duration)),
            ('Type de volontariat', lambda m: _label('volunteering_type', m.volunteering_type)),
            ('Activité', lambda m: m.activity_id.code_activite or ''),
            ('Nom légal organisme accueil',
             lambda m: m.hosting_partner_id.nom_legal or m.hosting_partner_id.name or ''),
            ("Pays d'accueil", lambda m: m.country_id.name or ''),
            ("Lieu de l'activité", lambda m: m.city or ''),
            ('Jeune', lambda m: _oui_non(m.jeune)),
            ('PRN', lambda m: m.prn or ''),
            ("ID de l'offre", lambda m: m.offer_code or ''),
            ("Intitulé de l'offre", lambda m: m.titre_offre or ''),
            ('Email du participant', lambda m: m.email or ''),
            ('Sexe du participant', lambda m: _label('sexe', m.sexe)),
            ('Âge du participant', lambda m: m.age),
            ('Nom légal organisme de soutien',
             lambda m: (_soutien_partner(m).nom_legal or _soutien_partner(m).name)
             if _soutien_partner(m) else ''),
            ('OID organisme de soutien',
             lambda m: _soutien_partner(m).oid or '' if _soutien_partner(m) else ''),
            ("Pays d'origine", lambda m: m.pays_residence_id.name or ''),
            ("Ville d'origine", lambda m: m.ville_residence or ''),
            ('Force majeure', lambda m: _oui_non(m.force_majeure)),
            ('Explications force majeure', lambda m: m.force_majeure_explication or ''),
            ('Soutien linguistique', lambda m: m.soutien_linguistique_type or ''),
            ('Tranche kilométrique', lambda m: m.tranche_kilometrique or ''),
            ('Distance réelle en kilomètres', lambda m: m.distance_reelle or 0),
            ('Moyen de transport principal', lambda m: m.transport_principal or ''),
            ('Voyage vert', lambda m: _oui_non(m.voyage_vert)),
            ('Date de début', lambda m: _date(m.start_date)),
            ('Date de fin', lambda m: _date(m.end_date)),
            ('Durée (jours)', lambda m: m.duree_jours),
            ('Durée SO (jours)', lambda m: m._get_finance_bm_summary()['soutien_organisationnel']['nb_jours']),
            ('Subvention SO par jour', lambda m: m._get_finance_bm_summary()['soutien_organisationnel']['montant_journalier']),
            ('Total SO', lambda m: m._get_finance_bm_summary()['soutien_organisationnel']['total']),
            ('Durée JAMO (jours)', lambda m: m._get_finance_bm_summary()['soutien_inclusion']['nb_jours']),
            ('Subvention JAMO par jour', lambda m: m._get_finance_bm_summary()['soutien_inclusion']['montant_journalier']),
            ('Total JAMO', lambda m: m._get_finance_bm_summary()['soutien_inclusion']['total']),
            ('Durée ADP (jours)', lambda m: m._get_finance_bm_summary()['argent_poche']['nb_jours']),
            ('Subvention ADP par jour', lambda m: m._get_finance_bm_summary()['argent_poche']['montant_journalier']),
            ('Total ADP', lambda m: m._get_finance_bm_summary()['argent_poche']['total']),
            ('Coûts exceptionnels', lambda m: m._get_finance_bm_summary()['cout_exceptionnel']['total']),
            ('Total de la subvention', lambda m: m._get_finance_bm_summary()['total_general']),
            ('Rapport demandé le', lambda m: _date(m.rapport_demande_le)),
            ('Rapport reçu le', lambda m: _date(m.rapport_recu_le)),
        ]
