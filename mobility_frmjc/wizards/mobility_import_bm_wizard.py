import base64
import datetime
import io
import logging

from odoo import models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

PROGRAMME_MAP = {
    'Service Civique (SC)': 'sc',
    'Corps Européen de Solidarité (CES)': 'ces',
    'Volontariat Service International (VSI)': 'vsi',
}
DIRECTION_MAP = {'Accueil': 'accueil', 'Envoi': 'envoi'}


def _parse_bm_date(value):
    if not value:
        return False
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.datetime.strptime(str(value), '%d/%m/%Y').date()
    except ValueError:
        return False


class MobilityImportBmWizard(models.TransientModel):
    _name = 'mobility.import.bm.wizard'
    _description = 'Assistant import BM (Beneficiary Module)'

    file = fields.Binary(string='Fichier BM (.xlsx)', required=True)
    filename = fields.Char()

    def action_importer(self):
        """Relit un export BM (§12 du cahier) et crée/met à jour les
        mobilités correspondantes.

        Clés de rapprochement : PRN → mobilité, code_activite → activité
        (§18 du cahier). Le rapprochement par OID sur les structures
        partenaires n'est PAS automatisé ici — même principe « humain
        dans la boucle » que le formulaire public de la Phase 6 : on
        évite de créer/modifier des fiches partenaires depuis un import
        de fichier sans revue. Le rattachement des structures reste une
        action manuelle du secrétariat après import.

        Pour la création de mobilités absentes, le participant est
        retrouvé par email (colonne « Email du participant ») ; sans
        correspondance, la ligne est ignorée plutôt que de créer un
        dossier incomplet.
        """
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise UserError("La bibliothèque openpyxl est requise. pip install openpyxl")

        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(self.file)), data_only=True,
        )
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

        def col(name):
            return headers.index(name) if name in headers else None

        idx_prn = col('PRN')
        idx_activite = col('Activité')
        idx_id_lieu = col('ID du lieu')
        idx_email = col('Email du participant')
        idx_programme = col('Programme')
        idx_direction = col('Type mobilité')
        idx_debut = col('Date de début')
        idx_fin = col('Date de fin')

        if idx_prn is None:
            raise UserError(
                'Colonne « PRN » introuvable — ce fichier ne semble pas '
                'être un export BM valide.'
            )

        Mobility = self.env['mobility.mobility']
        Activity = self.env['mobility.activity']
        Partner = self.env['res.partner']

        nb_crees = nb_maj = nb_ignores = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            prn = row[idx_prn] if idx_prn is not None else None
            if not prn:
                nb_ignores += 1
                continue

            vals = {'prn': prn}
            if idx_activite is not None and row[idx_activite]:
                activity = Activity.search(
                    [('code_activite', '=', row[idx_activite])], limit=1,
                )
                if activity:
                    vals['activity_id'] = activity.id
                    # Complète l'ID lieu s'il manque encore sur l'activité
                    # partagée — ne l'écrase jamais si déjà renseigné.
                    if (not activity.id_lieu and idx_id_lieu is not None
                            and row[idx_id_lieu]):
                        activity.id_lieu = row[idx_id_lieu]

            mobility = Mobility.search([('prn', '=', prn)], limit=1)
            if mobility:
                mobility.write(vals)
                nb_maj += 1
                continue

            email = row[idx_email] if idx_email is not None else None
            partner = Partner.search([('email', '=', email)], limit=1) if email else False
            if not partner:
                nb_ignores += 1
                continue

            vals.update({
                'participant_id': partner.id,
                'programme': (
                    PROGRAMME_MAP.get(row[idx_programme])
                    if idx_programme is not None else False
                ),
                'mobility_direction': (
                    DIRECTION_MAP.get(row[idx_direction])
                    if idx_direction is not None else False
                ),
                'start_date': _parse_bm_date(row[idx_debut]) if idx_debut is not None else False,
                'end_date': _parse_bm_date(row[idx_fin]) if idx_fin is not None else False,
            })
            try:
                Mobility.create(vals)
                nb_crees += 1
            except Exception:
                _logger.warning(
                    'Import BM : création impossible pour PRN %s', prn, exc_info=True,
                )
                nb_ignores += 1

        message = (
            f'{nb_crees} mobilité(s) créée(s), {nb_maj} mise(s) à jour, '
            f'{nb_ignores} ligne(s) ignorée(s) (PRN vide, ou participant '
            'introuvable pour une création).'
        )
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Import BM terminé',
                'message': message,
                'type': 'info',
                'sticky': True,
            },
        }
